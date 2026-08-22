#!/usr/bin/env python3
"""
Extrae los datos de una tabla desde un dump SQL (mysqldump) y los exporta a Excel.

Uso:
  python sql_to_excel.py                           # Usa tabla por defecto: inventario_inventario
  python sql_to_excel.py nombre_tabla              # Extrae la tabla indicada
  python sql_to_excel.py nombre_tabla salida.xlsx  # Archivo de salida personalizado
"""

import re
import sys
import ast
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("[!] Instalando openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
    import openpyxl


SQL_FILE = Path(__file__).parent / "data" / "localhost.sql"


def extract_columns(sql_text: str, table_name: str) -> list[str]:
    """Extrae los nombres de columnas del CREATE TABLE."""
    pattern = rf"CREATE TABLE `{re.escape(table_name)}`\s*\((.*?)\)\s*ENGINE"
    match = re.search(pattern, sql_text, re.DOTALL | re.IGNORECASE)
    if not match:
        return []

    body = match.group(1)
    columns = []
    for line in body.split('\n'):
        line = line.strip()
        col_match = re.match(r'^`(\w+)`', line)
        if col_match:
            columns.append(col_match.group(1))
    return columns


def extract_inserts(sql_text: str, table_name: str) -> list[str]:
    """Extrae todas las sentencias INSERT INTO para la tabla."""
    pattern = rf"INSERT INTO `{re.escape(table_name)}`.*?VALUES\s*"
    parts = re.split(pattern, sql_text, flags=re.IGNORECASE)
    if len(parts) < 2:
        return []
    # parts[1:] contiene los datos despues de cada INSERT INTO
    return parts[1:]


def parse_values(values_text: str) -> list[list]:
    """Parsea los VALUES de un INSERT INTO en filas."""
    rows = []
    # Buscar cada grupo (...) separado por comas
    depth = 0
    current = []
    in_string = False
    escape_next = False
    buf = ''

    for char in values_text:
        if escape_next:
            buf += char
            escape_next = False
            continue

        if char == '\\':
            buf += char
            escape_next = True
            continue

        if char == "'" and not escape_next:
            in_string = not in_string
            buf += char
            continue

        if in_string:
            buf += char
            continue

        if char == '(':
            depth += 1
            if depth == 1:
                buf = ''
                continue
        elif char == ')':
            depth -= 1
            if depth == 0:
                # Fin de una fila
                row = parse_row(buf)
                if row:
                    rows.append(row)
                buf = ''
                continue
        elif char == ';' and depth == 0:
            break

        if depth >= 1:
            buf += char

    return rows


def parse_row(row_str: str) -> list:
    """Parsea una fila de valores SQL separados por coma."""
    values = []
    current = ''
    in_string = False
    escape_next = False

    for char in row_str:
        if escape_next:
            current += char
            escape_next = False
            continue

        if char == '\\':
            escape_next = True
            continue

        if char == "'" and not escape_next:
            in_string = not in_string
            continue

        if char == ',' and not in_string:
            values.append(clean_value(current.strip()))
            current = ''
            continue

        current += char

    # Ultimo valor
    values.append(clean_value(current.strip()))
    return values


def clean_value(val: str):
    """Limpia un valor SQL y lo convierte al tipo Python apropiado."""
    if val == 'NULL' or val == '':
        return None
    # Si es un numero
    try:
        if '.' in val:
            return float(val)
        return int(val)
    except ValueError:
        pass
    return val


def main():
    table_name = sys.argv[1] if len(sys.argv) > 1 else 'industriales_industriales'
    output_file = sys.argv[2] if len(sys.argv) > 2 else f"scripts/data/{table_name}.xlsx"

    if not SQL_FILE.exists():
        print(f"[!] Archivo SQL no encontrado: {SQL_FILE}")
        sys.exit(1)

    print(f"[*] Leyendo {SQL_FILE} ({SQL_FILE.stat().st_size / 1024 / 1024:.1f} MB)...")
    sql_text = SQL_FILE.read_text(encoding='utf-8', errors='replace')

    # Extraer columnas
    columns = extract_columns(sql_text, table_name)
    if not columns:
        print(f"[!] No se encontro CREATE TABLE para `{table_name}`")
        # Listar tablas disponibles
        tables = re.findall(r"CREATE TABLE `(\w+)`", sql_text)
        if tables:
            print(f"[*] Tablas disponibles ({len(tables)}):")
            for t in sorted(tables):
                print(f"    - {t}")
        sys.exit(1)

    print(f"[*] Tabla: {table_name}")
    print(f"[*] Columnas ({len(columns)}): {', '.join(columns)}")

    # Extraer datos
    insert_parts = extract_inserts(sql_text, table_name)
    if not insert_parts:
        print(f"[!] No se encontraron INSERT INTO para `{table_name}`")
        sys.exit(1)

    all_rows = []
    for part in insert_parts:
        rows = parse_values(part)
        all_rows.extend(rows)

    print(f"[*] Filas extraidas: {len(all_rows)}")

    if not all_rows:
        print("[!] No se encontraron datos.")
        sys.exit(1)

    # Crear Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = table_name[:31]  # Max 31 chars for sheet name

    # Header
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = openpyxl.styles.Font(bold=True)

    # Data
    for row_idx, row_data in enumerate(all_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            if col_idx <= len(columns):
                ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-width
    for col_idx, col_name in enumerate(columns, 1):
        max_len = len(col_name)
        for row in ws.iter_rows(min_row=2, max_row=min(100, len(all_rows) + 1), min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 2, 50)

    # Guardar
    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"\n[OK] Archivo generado: {output_path}")
    print(f"     {len(all_rows)} filas x {len(columns)} columnas")


if __name__ == '__main__':
    main()
